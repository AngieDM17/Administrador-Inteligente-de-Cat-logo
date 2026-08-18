"""Pruebas de lote_masivo.py. Offline, sin red: investigar_producto() y
ejecutar_pipeline() se inyectan como dobles (mock.patch.object sobre este
modulo, mismo criterio que test_orquestador.py/test_agente_investigador.py)
-- en NINGUN caso se llama al agente investigador real, a Alibaba/Playwright,
a ElevenLabs ni a WooCommerce. Se prueba:

- Resiliencia: un lote de 3 links con uno fallando a proposito procesa los
  3 (no se detiene en el que falla) y el Excel queda con los 3 estados
  correctos escritos en disco.
- El Excel se reescribe DESPUES DE CADA producto, no solo al final.
- Una sola instancia de navegador_alibaba.SesionAlibaba para todo el lote,
  reusada entre varios links de Alibaba (el CAPTCHA se resuelve una vez).
- _texto_estado: traduccion pura de un resultado de _procesar_producto al
  texto de la columna Estado (listo / listo con avisos / fallo).
"""

import json
import tempfile
import unittest
from pathlib import Path
from unittest import mock

import openpyxl

import lote_masivo


def _armar_excel(ruta: Path, links: list[str], con_columna_nota: bool = True) -> None:
    libro = openpyxl.Workbook()
    hoja = libro.active
    if con_columna_nota:
        hoja.append(["Link", "Nota", "Estado"])
    else:
        hoja.append(["Link"])
    for link in links:
        if con_columna_nota:
            hoja.append([link, "", ""])
        else:
            hoja.append([link])
    libro.save(ruta)


def _leer_estados(ruta: Path) -> list[str]:
    libro = openpyxl.load_workbook(ruta)
    hoja = libro.active
    columnas = lote_masivo._indices_columnas(hoja)
    columna_estado = columnas["estado"]
    columna_link = columnas["link"]
    estados = []
    for fila in range(2, hoja.max_row + 1):
        link = hoja.cell(row=fila, column=columna_link).value
        if link:
            estados.append(hoja.cell(row=fila, column=columna_estado).value)
    return estados


def _investigar_producto_falso(link: str, carpeta_destino, publicar_notificacion,
                                evento_continuar=None, sesion_alibaba=None) -> dict:
    """Simula investigar_producto(): si el link contiene 'falla', devuelve
    error (sin llegar a escribir ninguna ficha); si no, escribe una ficha
    minima valida y devuelve 'ficha_lista', igual que la funcion real."""
    if "falla" in link:
        return {"estado": "error", "motivo": "la fuente esta caida (forzado en el test)"}

    carpeta_destino = Path(carpeta_destino)
    carpeta_destino.mkdir(parents=True, exist_ok=True)
    codigo = link.rstrip("/").rsplit("/", 1)[-1]
    ficha = {
        "entrada_original": {"codigo_proveedor": codigo, "link_producto": link},
        "producto": {"nombre_propuesto": f"Producto de prueba {codigo}"},
    }
    ruta_ficha = carpeta_destino / f"ficha_investigada_{codigo}.json"
    ruta_ficha.write_text(json.dumps(ficha), encoding="utf-8")
    return {"estado": "ficha_lista", "ruta_ficha": ruta_ficha}


class PruebasResiliencia(unittest.TestCase):
    """Un producto que falla en cualquier paso NUNCA frena el lote -- se
    anota su fila y se sigue con el siguiente."""

    def test_procesa_los_tres_aunque_uno_falle_y_escribe_los_tres_estados(self):
        with tempfile.TemporaryDirectory() as carpeta:
            carpeta = Path(carpeta)
            ruta_excel = carpeta / "lote.xlsx"
            carpeta_investigaciones = carpeta / "investigaciones"
            links = [
                "https://tienda.com/producto/AAA",
                "https://tienda.com/producto/falla-BBB",
                "https://tienda.com/producto/CCC",
            ]
            _armar_excel(ruta_excel, links)

            llamadas_pipeline = []

            def _ejecutar_pipeline_falso(ruta_ficha, publicar_notificacion,
                                         produccion=False, indice_producto=0):
                llamadas_pipeline.append(Path(ruta_ficha))
                return {"estado": "publicado", "producto_id": 111, "motivos_revision": []}

            mensajes = []
            with mock.patch.object(
                lote_masivo.agente_investigador, "investigar_producto",
                side_effect=_investigar_producto_falso,
            ):
                with mock.patch.object(
                    lote_masivo.orquestador, "ejecutar_pipeline",
                    side_effect=_ejecutar_pipeline_falso,
                ):
                    resultado = lote_masivo.procesar_lote(
                        ruta_excel, carpeta_investigaciones, mensajes.append,
                    )

            # Los 3 se procesaron -- el fallo del medio no corto el lote.
            self.assertEqual(resultado["total"], 3)
            self.assertEqual(len(resultado["procesados"]), 3)
            self.assertEqual(len(llamadas_pipeline), 2)  # el que fallo nunca llega aca

            estados = [p["estado"] for p in resultado["procesados"]]
            self.assertEqual(estados[0], "✅ listo")
            self.assertIn("falló", estados[1])
            self.assertIn("la fuente esta caida", estados[1])
            self.assertEqual(estados[2], "✅ listo")

            # Los mismos 3 estados quedaron ESCRITOS EN DISCO en el Excel.
            estados_en_disco = _leer_estados(ruta_excel)
            self.assertEqual(estados_en_disco, estados)

    def test_producto_publicado_con_avisos_del_colador_queda_marcado(self):
        with tempfile.TemporaryDirectory() as carpeta:
            carpeta = Path(carpeta)
            ruta_excel = carpeta / "lote.xlsx"
            carpeta_investigaciones = carpeta / "investigaciones"
            _armar_excel(ruta_excel, ["https://tienda.com/producto/DDD"])

            def _ejecutar_pipeline_falso(ruta_ficha, publicar_notificacion,
                                         produccion=False, indice_producto=0):
                return {
                    "estado": "publicado", "producto_id": 5,
                    "motivos_revision": ["Falta confirmar la potencia del motor"],
                }

            with mock.patch.object(
                lote_masivo.agente_investigador, "investigar_producto",
                side_effect=_investigar_producto_falso,
            ):
                with mock.patch.object(
                    lote_masivo.orquestador, "ejecutar_pipeline",
                    side_effect=_ejecutar_pipeline_falso,
                ):
                    resultado = lote_masivo.procesar_lote(
                        ruta_excel, carpeta_investigaciones, lambda m: None,
                    )

            estado = resultado["procesados"][0]["estado"]
            self.assertTrue(estado.startswith("⚠️ listo, revisar:"))
            self.assertIn("Falta confirmar la potencia del motor", estado)


class PruebasGuardadoIncremental(unittest.TestCase):
    def test_el_excel_se_guarda_una_vez_por_producto_no_solo_al_final(self):
        with tempfile.TemporaryDirectory() as carpeta:
            carpeta = Path(carpeta)
            ruta_excel = carpeta / "lote.xlsx"
            carpeta_investigaciones = carpeta / "investigaciones"
            _armar_excel(ruta_excel, [
                "https://tienda.com/producto/AAA",
                "https://tienda.com/producto/BBB",
                "https://tienda.com/producto/CCC",
            ])

            with mock.patch.object(
                lote_masivo.agente_investigador, "investigar_producto",
                side_effect=_investigar_producto_falso,
            ):
                with mock.patch.object(
                    lote_masivo.orquestador, "ejecutar_pipeline",
                    return_value={"estado": "publicado", "motivos_revision": []},
                ):
                    with mock.patch.object(
                        lote_masivo, "_guardar_libro",
                        wraps=lote_masivo._guardar_libro,
                    ) as guardar_falso:
                        lote_masivo.procesar_lote(
                            ruta_excel, carpeta_investigaciones, lambda m: None,
                        )

            self.assertEqual(guardar_falso.call_count, 3)


class PruebasSesionAlibabaCompartida(unittest.TestCase):
    """Requisito central del lote nocturno: el CAPTCHA/login de Alibaba se
    resuelve UNA vez por lote, no una vez por producto -- eso exige una
    UNICA instancia de SesionAlibaba reusada entre todos los links de
    Alibaba del lote."""

    def test_una_sola_sesion_para_varios_links_de_alibaba(self):
        with tempfile.TemporaryDirectory() as carpeta:
            carpeta = Path(carpeta)
            ruta_excel = carpeta / "lote.xlsx"
            carpeta_investigaciones = carpeta / "investigaciones"
            _armar_excel(ruta_excel, [
                "https://www.alibaba.com/product-detail/AAA.html",
                "https://www.alibaba.com/product-detail/BBB.html",
            ])

            sesiones_pasadas = []

            def _investigar_registrando_sesion(link, carpeta_destino,
                                                publicar_notificacion,
                                                evento_continuar=None,
                                                sesion_alibaba=None):
                sesiones_pasadas.append(sesion_alibaba)
                return _investigar_producto_falso(
                    link, carpeta_destino, publicar_notificacion,
                )

            class _SesionAlibabaFalsa:
                instancias = 0

                def __init__(self, *args, **kwargs):
                    type(self).instancias += 1
                    self.cerrada = False

                def cerrar(self):
                    self.cerrada = True

            with mock.patch.object(
                lote_masivo, "navegador_alibaba",
            ) as navegador_alibaba_falso:
                navegador_alibaba_falso.SesionAlibaba = _SesionAlibabaFalsa
                with mock.patch.object(
                    lote_masivo.agente_investigador, "investigar_producto",
                    side_effect=_investigar_registrando_sesion,
                ):
                    with mock.patch.object(
                        lote_masivo.orquestador, "ejecutar_pipeline",
                        return_value={"estado": "publicado", "motivos_revision": []},
                    ):
                        lote_masivo.procesar_lote(
                            ruta_excel, carpeta_investigaciones, lambda m: None,
                        )

            # Una sola instancia creada...
            self.assertEqual(_SesionAlibabaFalsa.instancias, 1)
            # ... y es LA MISMA que recibieron los dos links.
            self.assertEqual(len(sesiones_pasadas), 2)
            self.assertIsNotNone(sesiones_pasadas[0])
            self.assertIs(sesiones_pasadas[0], sesiones_pasadas[1])
            # Se cerro al terminar el lote.
            self.assertTrue(sesiones_pasadas[0].cerrada)

    def test_lote_sin_ningun_link_de_alibaba_no_abre_ninguna_sesion(self):
        with tempfile.TemporaryDirectory() as carpeta:
            carpeta = Path(carpeta)
            ruta_excel = carpeta / "lote.xlsx"
            carpeta_investigaciones = carpeta / "investigaciones"
            _armar_excel(ruta_excel, ["https://tienda.com/producto/AAA"])

            with mock.patch.object(
                lote_masivo, "navegador_alibaba",
            ) as navegador_alibaba_falso:
                with mock.patch.object(
                    lote_masivo.agente_investigador, "investigar_producto",
                    side_effect=_investigar_producto_falso,
                ):
                    with mock.patch.object(
                        lote_masivo.orquestador, "ejecutar_pipeline",
                        return_value={"estado": "publicado", "motivos_revision": []},
                    ):
                        lote_masivo.procesar_lote(
                            ruta_excel, carpeta_investigaciones, lambda m: None,
                        )

            navegador_alibaba_falso.SesionAlibaba.assert_not_called()


class PruebasColumnas(unittest.TestCase):
    def test_agrega_columna_estado_si_falta(self):
        with tempfile.TemporaryDirectory() as carpeta:
            ruta_excel = Path(carpeta) / "lote.xlsx"
            _armar_excel(ruta_excel, ["https://tienda.com/producto/AAA"], con_columna_nota=False)
            libro = openpyxl.load_workbook(ruta_excel)
            hoja = libro.active
            columnas = lote_masivo._indices_columnas(hoja)
            self.assertIn("link", columnas)
            self.assertIn("estado", columnas)
            self.assertEqual(hoja.cell(row=1, column=columnas["estado"]).value, "Estado")

    def test_sin_columna_link_lanza_value_error(self):
        with tempfile.TemporaryDirectory() as carpeta:
            ruta_excel = Path(carpeta) / "lote.xlsx"
            libro = openpyxl.Workbook()
            libro.active.append(["Nota", "Estado"])
            libro.save(ruta_excel)
            libro_leido = openpyxl.load_workbook(ruta_excel)
            with self.assertRaises(ValueError):
                lote_masivo._indices_columnas(libro_leido.active)


class PruebasTextoEstado(unittest.TestCase):
    """_texto_estado: logica pura, traduce el resultado de un producto al
    texto de la columna Estado."""

    def test_publicado_sin_avisos(self):
        resultado = {"estado": "publicado", "motivos_revision": []}
        self.assertEqual(lote_masivo._texto_estado(resultado), "✅ listo")

    def test_publicado_con_avisos(self):
        resultado = {
            "estado": "publicado",
            "motivos_revision": ["Falta la potencia", "Sin dimensiones"],
        }
        texto = lote_masivo._texto_estado(resultado)
        self.assertTrue(texto.startswith("⚠️ listo, revisar:"))
        self.assertIn("Falta la potencia", texto)
        self.assertIn("Sin dimensiones", texto)

    def test_revisar_categoria_cuenta_como_fallo_del_producto(self):
        resultado = {
            "estado": "revisar",
            "motivos": ["La categoria 'Gimnasio X' no existe en la tienda."],
        }
        texto = lote_masivo._texto_estado(resultado)
        self.assertTrue(texto.startswith("❌ falló:"))
        self.assertIn("Gimnasio X", texto)

    def test_error_generico(self):
        resultado = {"estado": "error", "motivo": "ffmpeg no encontrado"}
        self.assertEqual(lote_masivo._texto_estado(resultado), "❌ falló: ffmpeg no encontrado")


if __name__ == "__main__":
    unittest.main()

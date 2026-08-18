"""Corredor de lote nocturno (10-ago-2026).

Angie ya revisa a mano cada link de Alibaba (video real confirmado en la
pagina) y arma una lista en Excel. Este modulo lee ese Excel y procesa
TODOS los links, uno detras de otro, sin que ella este presente -- pensado
para dejarlo corriendo una jornada de 8 horas.

No reimplementa NADA del pipeline: para cada link llama, en el mismo orden
que ya usa app.py para un solo link, a
`agente_investigador.investigar_producto()` y despues
`orquestador.ejecutar_pipeline()`. La unica pieza nueva de verdad es la
COLA (leer el Excel, encadenar producto tras producto, escribir el estado
de cada uno) mas la sesion de Alibaba COMPARTIDA para todo el lote (ver
mas abajo).

--- Formato del Excel -------------------------------------------------

Una fila por producto, con estas columnas en la fila 1 (el orden no
importa, se buscan por nombre, sin distinguir mayusculas/acentos):

  Link    -- obligatoria. El link de producto (Alibaba u otra fuente).
  Nota    -- opcional, texto libre de Angie. No se toca.
  Estado  -- la llena este modulo. Si el Excel no la trae, se agrega sola.

--- Resiliencia: un fallo NUNCA frena el lote ---------------------------

Cada producto se procesa en su propio try/except (`_procesar_producto`):
si investigar_producto o ejecutar_pipeline fallan en cualquier paso
(investigacion, colador, galeria, video, publicacion), se anota el motivo
en la columna Estado de ESA fila y se sigue con el siguiente link. El
Excel se reescribe en disco despues de CADA producto, no solo al final --
si el proceso se corta a mitad de camino, el archivo ya refleja lo que se
alcanzo a hacer.

--- Una sola sesion de Alibaba para todo el lote ------------------------

`agente_investigador._correr_agente` abre una `navegador_alibaba.
SesionAlibaba` (ventana visible) nueva por cada llamada cuando el link es
de Alibaba -- correcto para UN link, pero repetiria el CAPTCHA/login una
vez por producto en un lote de varias decenas. Aca se crea una UNICA
instancia de `SesionAlibaba`, la primera vez que aparece un link de
Alibaba en la lista, y se pasa (parametro `sesion_alibaba`) a CADA llamada
de investigar_producto que la necesite -- la pausa de login/CAPTCHA solo
ocurre en la primera navegacion real de esa instancia
(`SesionAlibaba._primera_navegacion`), asi que el resto de los productos
de Alibaba del lote la reusan ya autenticada. La sesion se cierra al
final del lote completo (o si nunca hizo falta abrirla, no se abre).

--- Verificacion ---------------------------------------------------------

Sin red: ver test_lote_masivo.py. investigar_producto y ejecutar_pipeline
se inyectan como dobles (mock.patch.object sobre este modulo, mismo
criterio que test_orquestador.py/test_agente_investigador.py) -- nunca se
llama al agente real, a Alibaba, a ElevenLabs ni a WooCommerce.
"""

from __future__ import annotations

import json
import threading
import uuid
from pathlib import Path
from typing import Callable

import openpyxl

import agente_investigador
import navegador_alibaba
import orquestador

Notificador = Callable[[str], None]

COLUMNA_LINK = "link"
COLUMNA_ESTADO = "estado"

ESTADO_LISTO = "✅ listo"


def _texto_estado(resultado: dict) -> str:
    """Traduce el resultado de _procesar_producto (mismas formas que
    devuelve orquestador.ejecutar_pipeline, mas 'error' de la investigacion)
    al texto que va en la columna Estado del Excel. Logica pura."""
    estado = resultado.get("estado")

    if estado == "publicado":
        motivos = [m for m in (resultado.get("motivos_revision") or []) if m]
        if motivos:
            return f"⚠️ listo, revisar: {'; '.join(motivos)}"
        return ESTADO_LISTO

    if estado == "revisar":
        # Unico checkpoint que sigue frenando ejecutar_pipeline (categoria
        # sin match en la tienda, ver orquestador.py) -- el lote nocturno no
        # tiene a nadie presente para corregirlo en el momento, asi que se
        # anota como fallo de ESE producto (no del lote entero) para que
        # Angie lo revise a mano despues.
        motivos = [m for m in (resultado.get("motivos") or []) if m]
        detalle = "; ".join(motivos) if motivos else "necesita revision manual"
        return f"❌ falló: {detalle}"

    motivo = resultado.get("motivo") or "motivo desconocido"
    return f"❌ falló: {motivo}"


def _indices_columnas(hoja) -> dict[str, int]:
    """Mapea encabezado (fila 1, columna 1-index) -> nombre normalizado
    (minusculas, sin espacios alrededor). Si falta la columna 'Estado', se
    agrega una nueva al final -- un Excel armado a mano sin esa columna
    igual funciona, la llena el sistema. Lanza ValueError si falta 'Link':
    sin eso no hay nada que procesar."""
    columnas: dict[str, int] = {}
    ultima_con_encabezado = 0
    for indice in range(1, hoja.max_column + 1):
        valor = hoja.cell(row=1, column=indice).value
        if valor is None:
            continue
        nombre = str(valor).strip().lower()
        if not nombre:
            continue
        columnas[nombre] = indice
        ultima_con_encabezado = max(ultima_con_encabezado, indice)

    if COLUMNA_LINK not in columnas:
        raise ValueError(
            "El Excel no trae una columna 'Link' en la primera fila."
        )
    if COLUMNA_ESTADO not in columnas:
        ultima_con_encabezado += 1
        hoja.cell(row=1, column=ultima_con_encabezado, value="Estado")
        columnas[COLUMNA_ESTADO] = ultima_con_encabezado
    return columnas


def _guardar_libro(libro, ruta_excel: Path) -> None:
    """Reescribe el Excel en disco. Funcion propia (no inline en
    procesar_lote) para que los tests puedan verificar CUANTAS veces se
    guarda (una por producto, no solo al final) sin tener que re-leer el
    archivo despues de cada fila."""
    libro.save(ruta_excel)


def _procesar_producto(link: str, carpeta_investigaciones: Path,
                        publicar_notificacion: Notificador,
                        evento_continuar: threading.Event,
                        sesion_alibaba, produccion: bool = False) -> dict:
    """Corre investigar_producto() -> ejecutar_pipeline() para UN link,
    exactamente el mismo orden que _correr_pipeline de app.py para el
    camino de un solo link. Si la investigacion no produce una ficha, no se
    llega a ejecutar_pipeline. Nunca lanza: cualquier excepcion que se
    escape (ultima red de seguridad, mismo criterio que app.py) se traduce
    a {'estado': 'error', 'motivo': ...} en vez de tumbar el lote entero.

    `produccion`: ver orquestador.ejecutar_pipeline -- se pasa tal cual,
    mismo valor para TODO el lote (no se elige tienda producto por
    producto)."""
    carpeta_temporal = carpeta_investigaciones / uuid.uuid4().hex
    try:
        resultado_investigacion = agente_investigador.investigar_producto(
            link, carpeta_temporal, publicar_notificacion,
            evento_continuar=evento_continuar, sesion_alibaba=sesion_alibaba,
        )
        if resultado_investigacion["estado"] != "ficha_lista":
            return {
                "estado": "error",
                "motivo": resultado_investigacion.get(
                    "motivo", "La investigacion no produjo una ficha.",
                ),
            }

        ruta_ficha = Path(resultado_investigacion["ruta_ficha"])
        ficha = json.loads(ruta_ficha.read_text(encoding="utf-8-sig"))
        carpeta_nueva = agente_investigador.renombrar_carpeta_investigacion(
            ruta_ficha.parent, ficha, publicar_notificacion,
        )
        ruta_ficha = carpeta_nueva / ruta_ficha.name

        return orquestador.ejecutar_pipeline(
            ruta_ficha, publicar_notificacion, produccion=produccion,
        )
    except Exception as error:
        motivo = f"Error inesperado procesando '{link}': {error}"
        publicar_notificacion(motivo)
        return {"estado": "error", "motivo": motivo}


def procesar_lote(ruta_excel: Path, carpeta_investigaciones: Path,
                   publicar_notificacion: Notificador,
                   evento_continuar: threading.Event | None = None,
                   produccion: bool = False) -> dict:
    """Punto de entrada del lote nocturno. Lee `ruta_excel` (columnas Link /
    Nota opcional / Estado) y procesa cada fila con un link, encadenando
    investigar_producto() -> ejecutar_pipeline() (ver _procesar_producto),
    uno detras de otro y sin pausas entre productos (salvo la pausa real de
    Alibaba, que solo ocurre una vez por lote). Reescribe el Excel en disco
    despues de CADA producto.

    `produccion` (default False): igual para todas las filas del lote --
    ver orquestador.ejecutar_pipeline. Decision explicita de quien dispara
    el lote, nunca implicita.

    `evento_continuar` es el mismo `threading.Event` que ya usa el camino
    de un solo link (app.py, `_Job.evento_continuar`) para la pausa de
    login/CAPTCHA de Alibaba -- aca se REUSA para todo el lote: la
    SesionAlibaba unica de esta corrida lo espera una sola vez. Si no se
    pasa (ej. tests, o un lote sin ningun link de Alibaba), se crea uno
    propio que nunca se setea desde afuera.

    Devuelve {"total": N, "procesados": [...]}, con un resumen por fila
    ({"fila", "link", "estado"}) -- para que quien llama (el endpoint
    /generar-lote) pueda reportar el resultado final sin releer el Excel.
    Nunca lanza: un producto que falla se anota y el lote sigue."""
    ruta_excel = Path(ruta_excel)
    carpeta_investigaciones = Path(carpeta_investigaciones)
    if evento_continuar is None:
        evento_continuar = threading.Event()

    libro = openpyxl.load_workbook(ruta_excel)
    hoja = libro.active
    columnas = _indices_columnas(hoja)
    columna_link = columnas[COLUMNA_LINK]
    columna_estado = columnas[COLUMNA_ESTADO]

    filas_con_link = []
    for fila in range(2, hoja.max_row + 1):
        valor = hoja.cell(row=fila, column=columna_link).value
        link = str(valor).strip() if valor is not None else ""
        if link:
            filas_con_link.append((fila, link))

    total = len(filas_con_link)
    procesados: list[dict] = []
    sesion_alibaba = None

    try:
        for numero, (fila, link) in enumerate(filas_con_link, start=1):
            publicar_notificacion(f"Producto {numero} de {total}: {link}")

            if agente_investigador.es_alibaba(link) and sesion_alibaba is None:
                sesion_alibaba = navegador_alibaba.SesionAlibaba(
                    publicar_notificacion, evento_continuar,
                )

            resultado = _procesar_producto(
                link, carpeta_investigaciones, publicar_notificacion,
                evento_continuar, sesion_alibaba, produccion=produccion,
            )
            texto_estado = _texto_estado(resultado)

            hoja.cell(row=fila, column=columna_estado, value=texto_estado)
            _guardar_libro(libro, ruta_excel)

            publicar_notificacion(f"Producto {numero} de {total}: {texto_estado}")
            procesados.append({"fila": fila, "link": link, "estado": texto_estado})
    finally:
        if sesion_alibaba is not None:
            sesion_alibaba.cerrar()

    publicar_notificacion(f"Lote terminado: {total} producto(s) procesado(s).")
    return {"total": total, "procesados": procesados}
